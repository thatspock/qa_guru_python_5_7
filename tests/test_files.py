import os
import time
import requests
import xlrd
from zipfile import ZipFile
from selene import browser
from selenium import webdriver
from pypdf import PdfReader
from openpyxl import load_workbook

download_folder = 'resources'
current_path = os.path.abspath(__file__)
directory_path = os.path.dirname(current_path)
resources_path = os.path.join(directory_path, '..', download_folder)
resources_path = os.path.normpath(resources_path)
print(resources_path)


# TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp
def test_download_file_with_browser():
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": resources_path,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(10)

    downloaded_file_path = os.path.join(resources_path, 'pytest-main.zip')
    assert os.path.exists(downloaded_file_path), f"The file {downloaded_file_path} was not downloaded"

    file_size = os.path.getsize(downloaded_file_path)
    assert file_size > 0, f"The file {downloaded_file_path} is empty"


# TODO сохранять и читать из tmp, использовать универсальный путь
def test_downloaded_file_size():
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'

    r = requests.get(url)
    downloaded_file_path = os.path.join(resources_path, 'selenium_logo.png')

    with open(downloaded_file_path, 'wb') as file:
        file.write(r.content)

    size = os.path.getsize(downloaded_file_path)

    assert size == 30803


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_pdf():
    pdf_file_path = os.path.join(resources_path, 'docs-pytest-org-en-latest.pdf')

    assert os.path.exists(pdf_file_path), f'Pdf file {pdf_file_path} was not found'

    reader = PdfReader(pdf_file_path)

    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()

    assert number_of_pages == 412
    assert page is not None, 'Page is empty'
    assert text is not None and len(text.strip()) > 0, 'Text is empty'

    print(f'\n1: {number_of_pages}')
    print(f'2: {page}')
    print(f'3: {text}')


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xls():
    xls_file_path = os.path.join(resources_path, 'file_example_XLS_10.xls')

    assert os.path.exists(xls_file_path), f'XLS file {xls_file_path} was not found'

    book = xlrd.open_workbook(xls_file_path)

    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']

    print(f'\nКоличество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')

    sheet = book.sheet_by_index(0)

    assert sheet.ncols == 8
    assert sheet.nrows == 10
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'

    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx():
    xlsx_file_path = os.path.join(resources_path, 'file_example_XLSX_50.xlsx')

    assert os.path.exists(xlsx_file_path), f'XLSX file {xlsx_file_path} was not found'

    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active

    assert sheet.cell(row=3, column=2).value == 'Mara'

    print(sheet.cell(row=3, column=2).value)


# TODO Заархивировать имеющиеся в resources различные типы файлов в один архив
def test_zip():
    zip_file_path = os.path.join(resources_path, 'final.zip')

    pdf_pytest = os.path.join(resources_path, 'docs-pytest-org-en-latest.pdf')
    file_xls = os.path.join(resources_path, 'file_example_XLS_10.xls')
    file_xlsx = os.path.join(resources_path, 'file_example_XLSX_50.xlsx')
    logo = os.path.join(resources_path, 'selenium_logo.png')

    list_files = [pdf_pytest, file_xls, file_xlsx, logo]

    with ZipFile(zip_file_path, 'w') as zipF:
        for file in list_files:
            zipF.write(file, os.path.relpath(file, resources_path))

    with ZipFile(zip_file_path) as zipR:
        for file in list_files:
            unzipped_path = zipR.namelist()[list_files.index(file)]
            assert unzipped_path == os.path.relpath(file, resources_path)
            assert os.path.getsize(os.path.join(resources_path, unzipped_path)) == os.path.getsize(file)
