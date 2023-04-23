from zipfile import ZipFile
import time
from selene import browser
from selenium import webdriver
import os
import requests
from pypdf import PdfReader
import xlrd
from openpyxl import load_workbook


# TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp
def test_download_file_with_browser():
    options = webdriver.ChromeOptions()
    project_root_path = os.path.dirname(__file__)
    download_folder = 'resources'
    prefs = {
        "download.default_directory": os.path.join(project_root_path, download_folder),
        "download.prompt_for_download": False
    }

    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(10)

    downloaded_file_path = os.path.join(download_folder, 'pytest-main.zip')
    assert os.path.exists(downloaded_file_path), f"The file {downloaded_file_path} was not downloaded"

    file_size = os.path.getsize(downloaded_file_path)
    assert file_size > 0, f"The file {downloaded_file_path} is empty"


# TODO сохранять и читать из tmp, использовать универсальный путь
def test_downloaded_file_size():
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    download_folder = 'resources'

    r = requests.get(url)
    downloaded_file_path = os.path.join(download_folder, 'selenium_logo.png')

    with open(downloaded_file_path, 'wb') as file:
        file.write(r.content)

    size = os.path.getsize(downloaded_file_path)

    assert size == 30803


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_pdf():
    download_folder = 'resources'
    pdf_file_path = os.path.join(download_folder, 'docs-pytest-org-en-latest.pdf')

    assert os.path.exists(pdf_file_path), f'Pdf file {pdf_file_path} was not found'

    reader = PdfReader(pdf_file_path)

    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()

    assert number_of_pages == 412
    assert page is not None, 'Page is empty'
    assert text is not None and len(text.strip()) > 0, 'Text is empty'

    print(f'\n1: {number_of_pages}')
    print(f'2: {page}')
    print(f'3: {text}')


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xls():
    download_folder = 'resources'
    xls_file_path = os.path.join(download_folder, 'file_example_XLS_10.xls')

    assert os.path.exists(xls_file_path), f'XLS file {xls_file_path} was not found'

    book = xlrd.open_workbook(xls_file_path)

    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']

    print(f'\nКоличество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')

    sheet = book.sheet_by_index(0)

    assert sheet.ncols == 8
    assert sheet.nrows == 10
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'

    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx():
    download_folder = 'resources'
    xlsx_file_path = os.path.join(download_folder, 'file_example_XLSX_50.xlsx')

    assert os.path.exists(xlsx_file_path), f'XLSX file {xlsx_file_path} was not found'

    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active

    assert sheet.cell(row=3, column=2).value == 'Mara'

    print(sheet.cell(row=3, column=2).value)


# TODO Заархивировать имеющиеся в resources различные типы файлов в один архив
def test_zip():
    download_folder = 'resources'
    zip_file_path = os.path.join(download_folder, 'final.zip')

    pdf_pytest = os.path.join(download_folder, 'docs-pytest-org-en-latest.pdf')
    file_xls = os.path.join(download_folder, 'file_example_XLS_10.xls')
    file_xlsx = os.path.join(download_folder, 'file_example_XLSX_50.xlsx')
    logo = os.path.join(download_folder, 'selenium_logo.png')

    list_files = [pdf_pytest, file_xls, file_xlsx, logo]

    with ZipFile(zip_file_path, 'w') as zipF:
        for file in list_files:
            zipF.write(file)

    with ZipFile(zip_file_path) as zipR:
        for file in list_files:
            unzipped_path = zipR.namelist()[list_files.index(file)]
            assert unzipped_path == file
            assert os.path.getsize(unzipped_path) == os.path.getsize(file)
