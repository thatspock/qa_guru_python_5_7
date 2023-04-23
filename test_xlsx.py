from openpyxl import load_workbook
import os.path


# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    download_folder = 'resources'
    xlsx_file_path = os.path.join(download_folder, 'file_example_XLSX_50.xlsx')

    assert os.path.exists(xlsx_file_path), f'XLSX file {xlsx_file_path} was not found'

    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active

    assert sheet.cell(row=3, column=2).value == 'Mara'

    print(sheet.cell(row=3, column=2).value)
